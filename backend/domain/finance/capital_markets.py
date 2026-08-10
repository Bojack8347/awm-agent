"""Shared capital-market assumptions and risk/return utilities.

This module is intentionally read-only. It centralizes the long-term capital
market assumptions currently sourced by the asset allocation model from ``CMA.xlsx``
so cashflow and planning tools use the same asset-class universe, expected
returns, expected volatilities, and expected correlation matrix.

The first implementation is behavior-preserving for cashflow: same workbook
tabs, same formulas, and same fallback behavior when allocation data is absent
or unrecognized.
"""

from __future__ import annotations

import json
import math
import os
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Tuple

import numpy as np
from openpyxl import load_workbook


EXPECTED_RR_SHEET = "Expected R&R"
EXPECTED_CORR_SHEET = "Expected Corr Matrix"
RISK_RETURN_FRONTIER_SOURCE = "asset_allocation_frontier_cache"


def _backend_root() -> Path:
    return Path(__file__).resolve().parents[2]


def default_cma_workbook_path() -> Path:
    """Return the default workbook path, honoring an explicit env override."""
    override = os.getenv("AWM_CMA_WORKBOOK_PATH", "").strip()
    if override:
        return Path(override).expanduser()
    return (
        _backend_root()
        / "advisor"
        / "quant_models"
        / "asset_allocation_model"
        / "SAA Model"
        / "Inputs"
        / "L1 Asset Allocation"
        / "CMA.xlsx"
    )


def default_frontier_table_path() -> Path:
    """Return the cached asset allocation frontier table path."""
    override = os.getenv("AWM_RISK_RETURN_FRONTIER_PATH", "").strip()
    if override:
        return Path(override).expanduser()
    return Path(__file__).resolve().parent / "data" / "risk_return_table.json"


def _clean_asset_name(value: Any) -> str:
    return str(value or "").strip()


def _to_float(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        return float(value.strip())
    raise ValueError(f"Expected numeric value, got {value!r}")


@dataclass(frozen=True)
class CapitalMarketAssumptions:
    """Long-term assumptions loaded from the asset allocation CMA workbook."""

    asset_classes: Tuple[str, ...]
    expected_returns: Dict[str, float]
    expected_volatilities: Dict[str, float]
    expected_correlation_matrix: np.ndarray
    source_path: str
    source_sheets: Dict[str, str]

    @property
    def covariance_matrix(self) -> np.ndarray:
        vols = np.array(
            [self.expected_volatilities[name] for name in self.asset_classes],
            dtype=float,
        )
        return np.outer(vols, vols) * self.expected_correlation_matrix

    def metadata(self) -> Dict[str, Any]:
        return {
            "source": "CMA workbook",
            "source_path": self.source_path,
            "source_sheets": dict(self.source_sheets),
            "asset_class_count": len(self.asset_classes),
        }


@lru_cache(maxsize=8)
def load_capital_market_assumptions(
    workbook_path: Optional[str] = None,
) -> CapitalMarketAssumptions:
    """Load expected return, expected volatility, and correlations from Excel.

    The read layout mirrors asset allocation Layer 1's ``data_processor.load_saa_data``:
    - ``Expected R&R`` row 1: asset class names
    - ``Expected R&R`` row 2: expected returns
    - ``Expected R&R`` row 3: expected volatilities
    - ``Expected Corr Matrix``: row/column labels plus numeric matrix
    """
    path = Path(workbook_path).expanduser() if workbook_path else default_cma_workbook_path()
    if not path.exists():
        raise FileNotFoundError(f"Capital market assumptions workbook not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        rr = wb[EXPECTED_RR_SHEET]
        corr = wb[EXPECTED_CORR_SHEET]

        asset_classes = tuple(
            _clean_asset_name(cell)
            for cell in next(rr.iter_rows(min_row=1, max_row=1, values_only=True))[1:]
            if _clean_asset_name(cell)
        )
        if not asset_classes:
            raise ValueError("Expected R&R sheet did not contain asset classes")

        returns_row = next(rr.iter_rows(min_row=2, max_row=2, values_only=True))[1 : len(asset_classes) + 1]
        vols_row = next(rr.iter_rows(min_row=3, max_row=3, values_only=True))[1 : len(asset_classes) + 1]
        expected_returns = {
            asset: _to_float(value) for asset, value in zip(asset_classes, returns_row)
        }
        expected_volatilities = {
            asset: _to_float(value) for asset, value in zip(asset_classes, vols_row)
        }

        corr_headers = [
            _clean_asset_name(cell)
            for cell in next(corr.iter_rows(min_row=1, max_row=1, values_only=True))[1 : len(asset_classes) + 1]
        ]
        if tuple(corr_headers) != asset_classes:
            raise ValueError(
                "Expected Corr Matrix column headers do not match Expected R&R asset classes"
            )

        matrix_rows: List[List[float]] = []
        for row_idx, row in enumerate(
            corr.iter_rows(min_row=2, max_row=len(asset_classes) + 1, values_only=True)
        ):
            row_label = _clean_asset_name(row[0])
            expected_label = asset_classes[row_idx]
            if row_label != expected_label:
                raise ValueError(
                    f"Expected Corr Matrix row {row_idx + 2} label {row_label!r} "
                    f"does not match {expected_label!r}"
                )
            matrix_rows.append([
                _to_float(value) for value in row[1 : len(asset_classes) + 1]
            ])

        matrix = np.array(matrix_rows, dtype=float)
        if matrix.shape != (len(asset_classes), len(asset_classes)):
            raise ValueError(
                f"Expected correlation matrix shape {matrix.shape} does not match "
                f"{len(asset_classes)} asset classes"
            )

        return CapitalMarketAssumptions(
            asset_classes=asset_classes,
            expected_returns=expected_returns,
            expected_volatilities=expected_volatilities,
            expected_correlation_matrix=matrix,
            source_path=str(path),
            source_sheets={
                "expected_returns_and_volatilities": EXPECTED_RR_SHEET,
                "expected_correlation_matrix": EXPECTED_CORR_SHEET,
            },
        )
    finally:
        wb.close()


def normalize_weights(weights: Mapping[str, Any]) -> Dict[str, float]:
    """Normalize non-negative weights to sum to 1."""
    clean: Dict[str, float] = {}
    for key, value in (weights or {}).items():
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            continue
        clean[str(key).strip()] = max(0.0, numeric)
    total = sum(clean.values())
    if total <= 0:
        return {}
    return {key: value / total for key, value in clean.items()}


def estimate_allocation_risk_return(
    weights: Mapping[str, Any],
    *,
    default_return: float = 0.0,
    default_volatility: float = 0.0,
    assumptions: Optional[CapitalMarketAssumptions] = None,
) -> Dict[str, Any]:
    """Estimate expected return and volatility from asset-class weights.

    Unknown asset classes preserve the old cashflow behavior: if at least one
    recognized class exists, recognized weights are scaled to 100%; if every
    class is unknown or the allocation is empty, the caller's defaults are
    returned.
    """
    cma = assumptions or load_capital_market_assumptions()
    normalized = normalize_weights(weights)
    if not normalized:
        return {
            "expected_return": float(default_return),
            "volatility": float(default_volatility),
            "normalized_allocation": {},
            "recognized_allocation": {},
            "unrecognized_allocation": {},
            "fallback_used": True,
            "fallback_reason": "empty_allocation",
            "assumptions": cma.metadata(),
        }

    index = {name: idx for idx, name in enumerate(cma.asset_classes)}
    vector = np.zeros(len(cma.asset_classes), dtype=float)
    unrecognized: Dict[str, float] = {}

    for name, weight in normalized.items():
        idx = index.get(name)
        if idx is None:
            unrecognized[name] = weight
            continue
        vector[idx] += weight

    recognized_sum = float(vector.sum())
    if recognized_sum <= 0:
        return {
            "expected_return": float(default_return),
            "volatility": float(default_volatility),
            "normalized_allocation": normalized,
            "recognized_allocation": {},
            "unrecognized_allocation": unrecognized,
            "fallback_used": True,
            "fallback_reason": "all_asset_classes_unrecognized",
            "assumptions": cma.metadata(),
        }

    # Preserve previous cashflow behavior: unrecognized weight is
    # redistributed proportionally across known asset classes.
    if recognized_sum < 1.0:
        vector *= 1.0 / recognized_sum

    returns = np.array([cma.expected_returns[name] for name in cma.asset_classes], dtype=float)
    expected_return = float(vector @ returns)
    variance = float(vector @ cma.covariance_matrix @ vector)
    volatility = math.sqrt(max(0.0, variance))

    recognized = {
        asset: float(vector[idx])
        for idx, asset in enumerate(cma.asset_classes)
        if vector[idx] > 0
    }
    return {
        "expected_return": expected_return,
        "volatility": volatility,
        "normalized_allocation": normalized,
        "recognized_allocation": recognized,
        "unrecognized_allocation": unrecognized,
        "fallback_used": False,
        "fallback_reason": None,
        "assumptions": cma.metadata(),
    }


def analyze_allocation_risk_contributions(
    weights: Mapping[str, Any],
    *,
    stress_scenarios: Optional[List[Mapping[str, Any]]] = None,
    drawdown_config: Optional[Mapping[str, Any]] = None,
    fee_drag_config: Optional[Mapping[str, Any]] = None,
    initial_investment: Optional[float] = None,
    assumptions: Optional[CapitalMarketAssumptions] = None,
) -> Dict[str, Any]:
    """Calculate risk contributions, shocks, and explicit path/fee scenarios."""

    cma = assumptions or load_capital_market_assumptions()
    estimate = estimate_allocation_risk_return(
        weights,
        assumptions=cma,
    )
    if estimate.get("fallback_used"):
        raise ValueError(
            f"Risk contribution unavailable: {estimate.get('fallback_reason')}"
        )
    recognized = estimate["recognized_allocation"]
    vector = np.array(
        [float(recognized.get(name, 0.0)) for name in cma.asset_classes],
        dtype=float,
    )
    covariance = cma.covariance_matrix
    covariance_times_weights = covariance @ vector
    variance = float(vector @ covariance_times_weights)
    volatility = math.sqrt(max(0.0, variance))
    if volatility <= 0:
        raise ValueError("Portfolio volatility must be positive")

    contributions = []
    for index, asset_class in enumerate(cma.asset_classes):
        weight = float(vector[index])
        if weight <= 0:
            continue
        marginal_volatility = float(covariance_times_weights[index] / volatility)
        component_volatility = weight * marginal_volatility
        contributions.append(
            {
                "asset_class": asset_class,
                "weight": weight,
                "marginal_volatility": marginal_volatility,
                "component_volatility": component_volatility,
                "percentage_of_total_variance": (
                    float(weight * covariance_times_weights[index] / variance)
                    if variance > 0
                    else 0.0
                ),
            }
        )
    contributions.sort(
        key=lambda item: abs(item["component_volatility"]),
        reverse=True,
    )

    scenarios = stress_scenarios or _default_stress_scenarios()
    stress_results = []
    for scenario in scenarios:
        if not isinstance(scenario, Mapping):
            continue
        name = str(scenario.get("name") or "").strip()
        shocks = scenario.get("asset_class_shocks")
        if not name or not isinstance(shocks, Mapping):
            continue
        normalized_shocks = {}
        unknown = []
        portfolio_return = 0.0
        for asset_class, raw_shock in shocks.items():
            asset_name = str(asset_class)
            try:
                shock = float(raw_shock)
            except (TypeError, ValueError):
                raise ValueError(
                    f"Stress shock for {asset_name} must be numeric"
                )
            if not math.isfinite(shock) or not -1.0 <= shock <= 1.0:
                raise ValueError(
                    f"Stress shock for {asset_name} must be between -1 and 1"
                )
            if asset_name not in cma.asset_classes:
                unknown.append(asset_name)
                continue
            normalized_shocks[asset_name] = shock
            portfolio_return += float(recognized.get(asset_name, 0.0)) * shock
        stress_results.append(
            {
                "name": name,
                "portfolio_return": portfolio_return,
                "asset_class_shocks": normalized_shocks,
                "unrecognized_shock_asset_classes": unknown,
            }
        )
    drawdown_analysis = (
        _simulate_portfolio_max_drawdown(
            expected_return=float(estimate["expected_return"]),
            volatility=volatility,
            config=drawdown_config,
        )
        if drawdown_config is not None
        else None
    )
    fee_drag_analysis = (
        _portfolio_fee_drag(
            expected_return=float(estimate["expected_return"]),
            config=fee_drag_config,
            initial_investment=initial_investment,
        )
        if fee_drag_config is not None
        else None
    )
    limitations = [
        "Risk contributions use long-term covariance assumptions and are estimates.",
        "Stress results are deterministic one-period shocks; no recovery path or probability is implied.",
        "The analysis does not model taxes, trading costs, liquidity, or security-specific risk.",
    ]
    if drawdown_analysis is not None:
        limitations.append(
            "Maximum-drawdown results are seeded synthetic monthly paths under a "
            "constant-weight, monthly-rebalanced normal-return model; they are not "
            "historical backtests or guarantees."
        )
    if fee_drag_analysis is not None:
        limitations.append(
            "Fee drag uses the explicitly supplied blended annual fee, not verified "
            "security-level expense ratios, and excludes taxes and trading costs."
        )
    output = {
        "schema_version": "awm.portfolio_risk_analysis.v1",
        "expected_return_annual_decimal": estimate["expected_return"],
        "expected_volatility_annual_decimal": volatility,
        "risk_contributions": contributions,
        "stress_scenarios": stress_results,
        "assumptions": cma.metadata(),
        "methodology": {
            "risk_contribution": "component volatility = weight * (covariance @ weights) / portfolio volatility",
            "stress": "one-period weighted sum of explicit asset-class shocks; no recovery path or probability is implied",
            "drawdown": (
                "seeded monthly portfolio-return paths from the CMA annual mean and "
                "volatility, scaled to monthly values, with constant weights and "
                "monthly rebalancing"
                if drawdown_analysis is not None
                else None
            ),
            "fee_drag": (
                "annual compounding with the supplied blended fee subtracted from "
                "the CMA gross expected annual return"
                if fee_drag_analysis is not None
                else None
            ),
        },
        "limitations": limitations,
    }
    if drawdown_analysis is not None:
        output["drawdown_analysis"] = drawdown_analysis
    if fee_drag_analysis is not None:
        output["fee_drag_analysis"] = fee_drag_analysis
    return output


def _simulate_portfolio_max_drawdown(
    *,
    expected_return: float,
    volatility: float,
    config: Mapping[str, Any],
) -> Dict[str, Any]:
    """Return a reproducible distribution of path-based maximum drawdown."""

    if not isinstance(config, Mapping):
        raise ValueError("drawdown_config must be an object")
    horizon_years = _bounded_integer(
        config.get("horizon_years"),
        name="drawdown horizon_years",
        minimum=1,
        maximum=60,
    )
    num_simulations = _bounded_integer(
        config.get("num_simulations"),
        name="drawdown num_simulations",
        minimum=100,
        maximum=20_000,
    )
    seed = _bounded_integer(
        config.get("seed"),
        name="drawdown seed",
        minimum=0,
        maximum=2_147_483_647,
    )
    periods_per_year = 12
    period_count = horizon_years * periods_per_year
    monthly_mean = expected_return / periods_per_year
    monthly_volatility = volatility / math.sqrt(periods_per_year)
    generator = np.random.default_rng(seed)
    returns = generator.normal(
        loc=monthly_mean,
        scale=monthly_volatility,
        size=(num_simulations, period_count),
    )
    # A simple return cannot be below -100%. Clipping is disclosed and keeps
    # simulated wealth paths numerically and economically well-defined.
    returns = np.maximum(returns, -0.999999)
    wealth = np.cumprod(1.0 + returns, axis=1)
    wealth_with_opening = np.concatenate(
        [np.ones((num_simulations, 1), dtype=float), wealth],
        axis=1,
    )
    running_peak = np.maximum.accumulate(wealth_with_opening, axis=1)
    drawdowns = 1.0 - (wealth_with_opening / running_peak)
    maximum_drawdowns = np.max(drawdowns, axis=1)
    percentiles = np.percentile(maximum_drawdowns, [10, 50, 90])
    return {
        "type": "synthetic_maximum_drawdown_distribution",
        "maximum_drawdown_percentiles": {
            "p10": float(percentiles[0]),
            "p50": float(percentiles[1]),
            "p90": float(percentiles[2]),
        },
        "probability_maximum_drawdown_exceeds": {
            "10_percent": float(np.mean(maximum_drawdowns >= 0.10)),
            "20_percent": float(np.mean(maximum_drawdowns >= 0.20)),
            "30_percent": float(np.mean(maximum_drawdowns >= 0.30)),
        },
        "configuration": {
            "horizon_years": horizon_years,
            "num_simulations": num_simulations,
            "seed": seed,
            "periods_per_year": periods_per_year,
            "rebalancing": "monthly_constant_weight",
            "return_distribution": "normal",
            "simple_return_floor": -0.999999,
        },
        "inputs": {
            "expected_return_annual_decimal": expected_return,
            "expected_volatility_annual_decimal": volatility,
        },
    }


def _portfolio_fee_drag(
    *,
    expected_return: float,
    config: Mapping[str, Any],
    initial_investment: Optional[float],
) -> Dict[str, Any]:
    """Compound an explicitly supplied blended annual fee assumption."""

    if not isinstance(config, Mapping):
        raise ValueError("fee_drag_config must be an object")
    annual_fee_bps = _bounded_number(
        config.get("annual_fee_bps"),
        name="annual_fee_bps",
        minimum=0.0,
        maximum=1_000.0,
    )
    horizon_years = _bounded_integer(
        config.get("horizon_years"),
        name="fee horizon_years",
        minimum=1,
        maximum=60,
    )
    starting_value = (
        _bounded_number(
            initial_investment,
            name="initial_investment",
            minimum=0.01,
            maximum=1_000_000_000_000.0,
        )
        if initial_investment is not None
        else 100_000.0
    )
    fee_decimal = annual_fee_bps / 10_000.0
    net_expected_return = expected_return - fee_decimal
    if expected_return <= -1.0 or net_expected_return <= -1.0:
        raise ValueError("expected return after fee must be greater than -100%")
    gross_terminal = starting_value * ((1.0 + expected_return) ** horizon_years)
    net_terminal = starting_value * ((1.0 + net_expected_return) ** horizon_years)
    drag = gross_terminal - net_terminal
    return {
        "type": "blended_annual_fee_scenario",
        "annual_fee_bps": annual_fee_bps,
        "annual_fee_decimal": fee_decimal,
        "horizon_years": horizon_years,
        "initial_investment": starting_value,
        "initial_investment_source": (
            "signed_allocation_mandate"
            if initial_investment is not None
            else "standardized_100000_dollar_basis"
        ),
        "gross_expected_return_annual_decimal": expected_return,
        "net_expected_return_annual_decimal": net_expected_return,
        "gross_terminal_value": gross_terminal,
        "net_terminal_value_after_fee": net_terminal,
        "cumulative_fee_drag": drag,
        "cumulative_fee_drag_as_share_of_gross_terminal": (
            drag / gross_terminal if gross_terminal else 0.0
        ),
        "fee_assumption_source": "explicit_tool_input_not_security_level_lookup",
    }


def _bounded_integer(
    value: Any,
    *,
    name: str,
    minimum: int,
    maximum: int,
) -> int:
    if isinstance(value, bool) or not isinstance(value, int):
        raise ValueError(f"{name} must be an integer")
    if not minimum <= value <= maximum:
        raise ValueError(f"{name} must be between {minimum} and {maximum}")
    return value


def _bounded_number(
    value: Any,
    *,
    name: str,
    minimum: float,
    maximum: float,
) -> float:
    if isinstance(value, bool):
        raise ValueError(f"{name} must be numeric")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{name} must be numeric") from exc
    if not math.isfinite(number) or not minimum <= number <= maximum:
        raise ValueError(f"{name} must be between {minimum} and {maximum}")
    return number


def _default_stress_scenarios() -> List[Dict[str, Any]]:
    return [
        {
            "name": "equity_drawdown",
            "asset_class_shocks": {
                "US Equity": -0.30,
                "Dev. Europe ex UK Equity": -0.30,
                "Japan Equity": -0.28,
                "China Equity": -0.35,
                "India Equity": -0.30,
                "Global High Yield Bond BB-B": -0.12,
                "Emerging Market Local Currency Government Bonds": -0.15,
                "Emerging Market Hard Currency Debt": -0.15,
                "Commodities": -0.12,
                "Hedge Funds": -0.10,
                "Bitcoin": -0.55,
                "Gold": 0.05,
                "US Treasury": 0.06,
                "Global Investment Grade Corporate Bond": -0.03,
                "Cash": 0.0,
            },
        },
        {
            "name": "rates_up",
            "asset_class_shocks": {
                "US Treasury": -0.08,
                "Global Investment Grade Corporate Bond": -0.10,
                "Global High Yield Bond BB-B": -0.08,
                "Emerging Market Local Currency Government Bonds": -0.10,
                "Emerging Market Hard Currency Debt": -0.10,
                "US Equity": -0.08,
                "Dev. Europe ex UK Equity": -0.08,
                "Japan Equity": -0.08,
                "China Equity": -0.10,
                "India Equity": -0.08,
                "Commodities": 0.05,
                "Gold": -0.06,
                "Hedge Funds": -0.04,
                "Bitcoin": -0.15,
                "Cash": 0.02,
            },
        },
    ]


def allocation_assumptions(
    weights: Mapping[str, Any],
    default_return: float,
    default_volatility: float,
    *,
    assumptions: Optional[CapitalMarketAssumptions] = None,
) -> Tuple[float, float]:
    """Compatibility wrapper returning only ``(expected_return, volatility)``."""
    estimate = estimate_allocation_risk_return(
        weights,
        default_return=default_return,
        default_volatility=default_volatility,
        assumptions=assumptions,
    )
    return float(estimate["expected_return"]), float(estimate["volatility"])


def blend_account_pool_allocation(pool: Any) -> Dict[str, float]:
    """Balance-weight allocation across an account pool.

    This preserves the cashflow API's prior semantics exactly, including the
    behavior where accounts with empty allocation dictionaries contribute to
    the denominator before the final allocation is normalized by the estimator.
    """
    if not isinstance(pool, list):
        return {}

    total_balance = 0.0
    weighted_alloc: Dict[str, float] = {}
    for acct in pool:
        if not isinstance(acct, dict):
            continue
        try:
            balance = float(acct.get("balance", 0.0) or 0.0)
        except (TypeError, ValueError):
            balance = 0.0
        allocation = acct.get("allocation")
        if not isinstance(allocation, dict) or balance <= 0:
            continue
        acct_norm = normalize_weights(allocation)
        for cls_name, weight in acct_norm.items():
            weighted_alloc[cls_name] = weighted_alloc.get(cls_name, 0.0) + weight * balance
        total_balance += balance

    if total_balance <= 0:
        return {}
    return {key: value / total_balance for key, value in weighted_alloc.items()}


def estimate_account_pool_assumptions(
    pool: Any,
    *,
    default_return: float = 0.0,
    default_volatility: float = 0.0,
    assumptions: Optional[CapitalMarketAssumptions] = None,
) -> Dict[str, Any]:
    """Estimate risk/return for a balance-weighted account pool."""
    blended = blend_account_pool_allocation(pool)
    result = estimate_allocation_risk_return(
        blended,
        default_return=default_return,
        default_volatility=default_volatility,
        assumptions=assumptions,
    )
    result["blended_allocation"] = blended
    return result


def load_risk_return_frontier_table(path: Optional[str] = None) -> Dict[str, Any]:
    table_path = Path(path).expanduser() if path else default_frontier_table_path()
    with table_path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def lookup_risk_return_frontier(
    *,
    required_return_pct: Optional[float] = None,
    target_volatility_pct: Optional[float] = None,
    path: Optional[str] = None,
) -> Dict[str, Any]:
    """Lookup the cached asset allocation risk-return frontier.

    If ``target_volatility_pct`` is provided, returns the nearest volatility
    row. If ``required_return_pct`` is provided, returns the first row whose
    expected return meets or exceeds that requirement, plus nearest lower and
    upper frontier rows for explanation.
    """
    table = load_risk_return_frontier_table(path)
    entries = [
        entry
        for entry in table.get("entries", [])
        if isinstance(entry, dict) and entry.get("expected_return_pct") is not None
    ]
    entries.sort(key=lambda entry: float(entry.get("target_volatility_pct", 0.0)))
    if not entries:
        return {
            "success": False,
            "error": "risk-return frontier table has no usable entries",
            "source": RISK_RETURN_FRONTIER_SOURCE,
        }

    selected = None
    lower = None
    upper = None
    feasible = True
    mode = None

    if target_volatility_pct is not None:
        mode = "target_volatility"
        target = float(target_volatility_pct)
        selected = min(
            entries,
            key=lambda entry: abs(float(entry.get("target_volatility_pct", 0.0)) - target),
        )
    elif required_return_pct is not None:
        mode = "required_return"
        required = float(required_return_pct)
        for entry in entries:
            expected = float(entry.get("expected_return_pct", 0.0))
            if expected < required:
                lower = entry
                continue
            upper = entry
            selected = entry
            break
        if selected is None:
            feasible = False
            selected = entries[-1]
            lower = entries[-1]
        if upper is None and feasible:
            upper = selected
    else:
        return {
            "success": False,
            "error": "required_return_pct or target_volatility_pct is required",
            "source": RISK_RETURN_FRONTIER_SOURCE,
        }

    return {
        "success": True,
        "mode": mode,
        "required_return_pct": required_return_pct,
        "target_volatility_pct": target_volatility_pct,
        "feasible": feasible,
        "selected_entry": selected,
        "lower_entry": lower,
        "upper_entry": upper,
        "source": RISK_RETURN_FRONTIER_SOURCE,
        "frontier_metadata": {
            "description": table.get("description"),
            "generated_at": table.get("generated_at"),
            "parameters": table.get("parameters", {}),
            "entry_count": len(entries),
        },
        "assumptions": load_capital_market_assumptions().metadata(),
    }
